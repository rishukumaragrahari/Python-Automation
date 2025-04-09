# -*- coding: utf-8 -*-
"""
Created on Fri Sep 15 23:48:16 2022

@author: Rishav Kumar Agrahari
"""

# Call all required modules
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from openpyxl.styles import Border, Side
import time
from datetime import datetime
from datetime import date
import numpy as np


# condition to Check system version and use library according to system version
if sys.version_info[0] < 3: 
   	from StringIO import StringIO
else:
   	from io import StringIO
       
# Define input & output path

INPATH =r"C:\\Users\\edebkun\\OneDrive - Ericsson\\python\\Claro\\SSV_Automations\\WCDMA\\Input\\"
OUTPATH =r"C:\\Users\\edebkun\\OneDrive - Ericsson\\python\\Claro\\SSV_Automations\\WCDMA\\Output\\"



############### Reading Input Details #################

df_input = pd.read_excel(INPATH+'Input_Details.xlsx', sheet_name="Input")

typeofnode=df_input.at[0,'Data']
nodename=df_input.at[1,'Data']
rnc_name=df_input.at[2,'Data']

cell_list=list(df_input["Data"].loc[3:])


if typeofnode=="Baseband":
    # Function to get data b/w two strings
    
    def find_between( s, first, last ):
        try:
            start = s.index( first ) -12
            end = s.index( last, start )
            return s[start:end]
        except ValueError:
            return ""
    
    def find_between_rnc_commands( s, first, last ):
        try:
            start = s.index( first )
            end = s.index( last, start )
            return s[start:end]
        except ValueError:
            return ""
    
    def get_line_with_substring(input_string, substring):
        lines = input_string.split('\n')
        
        for line in lines:
            if substring in line:
                return line
        return None
        
    # Define border style for excel
    border = Border(
        left=Side(style="thick"),
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick")
        )
    
    def apply_border( cell_range ):
        # Split the cell range into top-left and bottom-right coordinates
        top_left_cell, bottom_right_cell = cell_range.split(":")
        
        # Set the outline border for the range
        min_row, min_col = ws[top_left_cell].row, ws[top_left_cell].column
        max_row, max_col = ws[bottom_right_cell].row, ws[bottom_right_cell].column
        
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if row == min_row:
                    cell.border = cell.border + Border(top=border.top)
                if row == max_row:
                    cell.border = cell.border + Border(bottom=border.bottom)
                if col == min_col:
                    cell.border = cell.border + Border(left=border.left)
                if col == max_col:
                    cell.border = cell.border + Border(right=border.right)
                
    
    ######################################### Baseband SSV CHECKLIST ##############################################################################
    
    ##################    Loading WCDMA BB  ###########################  
    print("Loading Log...")
    with open (INPATH + os.sep+ 'BB_Log.log', "r") as file:
    	data = file.read().rstrip() 
    
    data = data.replace("|", "_")
    
    ##################    Loading Workbook  ###########################    
    
    wb = load_workbook(INPATH +os.sep+'3G_Checklist_Template.xlsx')  
    
    ##################    VSWR  ########################### 
    
    print("Getting VSWR")    
       
    start_string="invxrf"
    end_string="st ret"
    dict_replace={'invxrf':''}
    
    data_between=find_between( data, start_string, end_string )
    
    stringio_data_between= StringIO(data_between)
    df_data_between = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")
    
    df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)
    
    ws = wb['VSWR']
    
    rows = dataframe_to_rows(df_data_between,index=False,header=False)
    for r_idx, row in enumerate(rows, 4):
    	for c_idx, value in enumerate(row, 2):
    		if str(value)[0]=="=":
    			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
    		else:
    			ws.cell(row=r_idx, column=c_idx, value=str(value))
    
    cell_range = "B4:B"+str(len(df_data_between)+3)
    apply_border( cell_range )
    
    
    ##################    RET  ########################### 
    
    print("Getting RET")    
       
    start_string="st ret"
    end_string="get 0"
    dict_replace={'st ret':'','lget . Electricalantennatilt':''}
    
    data_between=find_between( data, start_string, end_string )
    
    stringio_data_between= StringIO(data_between)
    df_data_between = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")
    
    df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)
    
    ws = wb['RET']
    
    rows = dataframe_to_rows(df_data_between,index=False,header=False)
    for r_idx, row in enumerate(rows, 4):
    	for c_idx, value in enumerate(row, 2):
    		if str(value)[0]=="=":
    			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
    		else:
    			ws.cell(row=r_idx, column=c_idx, value=str(value))
    
    cell_range = "B4:B"+str(len(df_data_between)+3)
    apply_border( cell_range )
    
    
    ##################    EQUIPMENT  ########################### 
    
    print("Getting EQUIPMENT")    
       
    start_string="get 0"
    end_string="lget . ^keyId"
    dict_replace={'get 0':''}
    
    data_between=find_between( data, start_string, end_string )
    
    stringio_data_between= StringIO(data_between)
    df_data_between = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")
    
    df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)
    
    ws = wb['EQUIPMENT']
    
    rows = dataframe_to_rows(df_data_between,index=False,header=False)
    for r_idx, row in enumerate(rows, 4):
    	for c_idx, value in enumerate(row, 2):
    		if str(value)[0]=="=":
    			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
    		else:
    			ws.cell(row=r_idx, column=c_idx, value=str(value))
    
    cell_range = "B4:B"+str(len(df_data_between)+3)
    apply_border( cell_range )
    
    
    ##################    BRD  ########################### 
    
    print("Getting BRD")    
       
    start_string="invxrf"
    end_string="st ret"
    dict_replace={'invxrf':''}
    
    data_between=find_between( data, start_string, end_string )
    
    stringio_data_between= StringIO(data_between)
    df_data_between = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")
    
    df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)
    
    ws = wb['BRD']
    
    rows = dataframe_to_rows(df_data_between,index=False,header=False)
    for r_idx, row in enumerate(rows, 4):
    	for c_idx, value in enumerate(row, 2):
    		if str(value)[0]=="=":
    			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
    		else:
    			ws.cell(row=r_idx, column=c_idx, value=str(value))
    
    cell_range = "B4:B"+str(len(df_data_between)+3)
    apply_border( cell_range )
    
    
    ##################    LICENSE  ########################### 
    
    print("Getting LICENSE")    
       
    start_string="lget . ^keyId"
    end_string="> pst"
    dict_replace={'lget . ^keyId':''}
    
    data_between=find_between( data, start_string, end_string )
    
    stringio_data_between= StringIO(data_between)
    df_data_between = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")
    
    df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)
    
    ws = wb['LICENSE']
    
    rows = dataframe_to_rows(df_data_between,index=False,header=False)
    for r_idx, row in enumerate(rows, 4):
    	for c_idx, value in enumerate(row, 2):
    		if str(value)[0]=="=":
    			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
    		else:
    			ws.cell(row=r_idx, column=c_idx, value=str(value))
    
    cell_range = "B4:B"+str(len(df_data_between)+3)
    apply_border( cell_range )
    
    
    ##################    COUNTER  ########################### 
    
    print("Getting COUNTER")    
       
    start_string="> pst"
    end_string="lget iub_"
    dict_replace={'pst':''}
    
    data_between=find_between( data, start_string, end_string )
    
    stringio_data_between= StringIO(data_between)
    df_data_between = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")
    
    df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)
    
    ws = wb['COUNTER']
    
    rows = dataframe_to_rows(df_data_between,index=False,header=False)
    for r_idx, row in enumerate(rows, 4):
    	for c_idx, value in enumerate(row, 2):
    		if str(value)[0]=="=":
    			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
    		else:
    			ws.cell(row=r_idx, column=c_idx, value=str(value))
    
    cell_range = "B4:B"+str(len(df_data_between)+3)
    apply_border( cell_range )
    
    
    
    ##################    E1  ########################### 
    
    print("Getting E1")    
       
    start_string="lget iub_"
    end_string="> alt"
    dict_replace={'lget iub_':''}
    
    data_between=find_between( data, start_string, end_string )
    
    stringio_data_between= StringIO(data_between)
    df_data_between = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")
    
    df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)
    
    ws = wb['E1']
    
    rows = dataframe_to_rows(df_data_between,index=False,header=False)
    for r_idx, row in enumerate(rows, 4):
    	for c_idx, value in enumerate(row, 2):
    		if str(value)[0]=="=":
    			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
    		else:
    			ws.cell(row=r_idx, column=c_idx, value=str(value))
    
    cell_range = "B4:B"+str(len(df_data_between)+3)
    apply_border( cell_range )
    
    
    ##################    ALARMS  ########################### 
    
    print("Getting ALARMS")    
       
    start_string="> alt"
    end_string="invl power"
    dict_replace={'alt':''}
    
    data_between=find_between( data, start_string, end_string )
    
    stringio_data_between= StringIO(data_between)
    df_data_between = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")
    
    df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)
    
    ws = wb['ALARMS']
    
    rows = dataframe_to_rows(df_data_between,index=False,header=False)
    for r_idx, row in enumerate(rows, 4):
    	for c_idx, value in enumerate(row, 2):
    		if str(value)[0]=="=":
    			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
    		else:
    			ws.cell(row=r_idx, column=c_idx, value=str(value))
    
    cell_range = "B4:B"+str(len(df_data_between)+3)
    apply_border( cell_range )
    
    
    
    ##################    LAC  ########################### 
    
    print("Getting LAC")    
       
    line=get_line_with_substring(data, "utrancellid_localcellid_uarfcndl_uarfcnul_locationAreaRef")
    
    start_string=line
    end_string="get loc lac"
    dict_replace={line[9:]:''}
    
    data_between=find_between_rnc_commands( data, start_string, end_string )
    
    stringio_data_between= StringIO(data_between)
    df_data_between = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")
    
    df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)
    
    ws = wb['LAC']
    
    rows = dataframe_to_rows(df_data_between,index=False,header=False)
    for r_idx, row in enumerate(rows, 4):
    	for c_idx, value in enumerate(row, 2):
    		if str(value)[0]=="=":
    			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
    		else:
    			ws.cell(row=r_idx, column=c_idx, value=str(value))
    
    cell_range = "B4:B"+str(len(df_data_between)+3)
    apply_border( cell_range )
    
    
    
    ##################    Working on NBRs  ########################### 
    
    print("Getting NBRs")   
    
    nbr_ws_dict={0:'NBRList_850A_S1',1:'NBRList_850A_S2',2:'NBRList_850A_S3',3:'NBRList_850A_S4',
                 4:'NBRList_850B_S1',5:'NBRList_850B_S2',6:'NBRList_850B_S3',7:'NBRList_850B_S4',
                 8:'NBRList_2100A_S1',9:'NBRList_2100A_S2',10:'NBRList_2100A_S3',11:'NBRList_2100A_S4',
                 12:'NBRList_2100B_S1',13:'NBRList_2100B_S2',14:'NBRList_2100B_S3',15:'NBRList_2100B_S4'}
    cell_count=0
    
    for cell in cell_list:
        if cell=="NO INFO":
            ws = wb[nbr_ws_dict[cell_count]]
            ws.sheet_state = 'hidden'
            cell_count=cell_count+1
        else:
            start_string="lget utrancell="+cell+' utranrelation_gsmrelation'
            end_string="MOs"
            dict_replace={start_string:''}
    
            data_between=find_between( data, start_string, end_string )
    
            stringio_data_between= StringIO(data_between)
            df_data_between = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")
    
            df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)
    
            ws = wb[nbr_ws_dict[cell_count]]
    
            rows = dataframe_to_rows(df_data_between,index=False,header=False)
            for r_idx, row in enumerate(rows, 4):
            	for c_idx, value in enumerate(row, 2):
            		if str(value)[0]=="=":
            			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
            		else:
            			ws.cell(row=r_idx, column=c_idx, value=str(value))
    
            cell_range = "B4:B"+str(len(df_data_between)+3)
            apply_border( cell_range )
            
            cell_count=cell_count+1
    
            
    ########### Saving 3G_Checklist ###########
    
    print("Saving 3G_Checklist.")  
    
    now = datetime.now()
    today = date.today()
    
    
    date_today = time.strftime('%b-%d-%Y', time.localtime())
    wb.save(OUTPATH+os.sep+'3G_Checklist_'+nodename+"_"+date_today+'.xlsx')
    
    
    
    ########### Working on power checklist ###########
    
    ##################    Loading Workbook  ###########################    
    
    wb = load_workbook(INPATH +os.sep+'3G_Checklist_Power_Template.xlsx')  
    
    
    ##################    POWER  ########################### 
    
    print("Getting POWER")    
       
    start_string="invl power"
    end_string="get . configuredMaxTxPower"
    dict_replace={'invl power':''}
    
    data_between=find_between( data, start_string, end_string )
    
    stringio_data_between= StringIO(data_between)
    df_data_between = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")
    
    df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)
    
    ws = wb['POWER']
    
    rows = dataframe_to_rows(df_data_between,index=False,header=False)
    for r_idx, row in enumerate(rows, 4):
    	for c_idx, value in enumerate(row, 2):
    		if str(value)[0]=="=":
    			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
    		else:
    			ws.cell(row=r_idx, column=c_idx, value=str(value))
    
    cell_range = "B4:B"+str(len(df_data_between)+3)
    apply_border( cell_range )
    
    
    ##################    MAX_POWER_CONFIG  ########################### 
    
    print("Getting MAX_POWER_CONFIG")    
       
    start_string="get . configuredMaxTxPower"
    end_string="exit"
    dict_replace={'get . configuredMaxTxPower':''}
    
    data_between=find_between( data, start_string, end_string )
    
    stringio_data_between= StringIO(data_between)
    df_data_between = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")
    
    df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)
    
    ws = wb['MAX_POWER_CONFIG']
    
    rows = dataframe_to_rows(df_data_between,index=False,header=False)
    for r_idx, row in enumerate(rows, 4):
    	for c_idx, value in enumerate(row, 2):
    		if str(value)[0]=="=":
    			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
    		else:
    			ws.cell(row=r_idx, column=c_idx, value=str(value))
    
    cell_range = "B4:B"+str(len(df_data_between)+3)
    apply_border( cell_range )
    
    
    
    ########### Saving Power CheckList ###########
    
    print("Saving Power CheckList.")  
    
    now = datetime.now()
    today = date.today()
    
    
    date_today = time.strftime('%b-%d-%Y', time.localtime())
    wb.save(OUTPATH+os.sep+'3G_Checklist_Power_'+nodename+"_"+date_today+'.xlsx')
    
    
    
    
    
    
    
    
    ######################## Working on SSV_FILE_INPUTS #######################
    
    
    print("Saving SSV_FILE_INPUTS.")
    
    line=get_line_with_substring(data, "cId_locationAreaRef_maximumTransmissionPower_primaryCpichPower_primaryScramblingCodee_uarfcnDl_uarfcnUl")
    
    start_string=line
    end_string="get loc lac"
    dict_replace={line[9:]:''}
     
    data_between=find_between_rnc_commands( data, start_string, end_string )
    
    
    stringio_data_between= StringIO(data_between)
    df_ssv_input = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")
    
    df_ssv_input=df_ssv_input.loc[(df_ssv_input[0].str.startswith('Utran', na=False))|(df_ssv_input[0].str.startswith('MO', na=False))]
    
    df_ssv_input=df_ssv_input[0].str.split('\s+', expand=True)
    
    df_ssv_input[0] = df_ssv_input[0].str.replace('UtranCell=','')
    
    df_ssv_input = df_ssv_input.reset_index(drop=True)
    
    new_header = df_ssv_input.iloc[0] 
    df_ssv_input = df_ssv_input[1:]
    df_ssv_input.columns = new_header
    
    df_ssv_input = df_ssv_input.reset_index(drop=True)
    
    print("Mapping LAC.")
    
    start_string="get loc lac"
    end_string="Total"
    dict_replace={start_string:''}
    
    data_between=find_between( data, start_string, end_string )
    
    stringio_data_between= StringIO(data_between)
    df_lac = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")
    
    
    df_lac=df_lac.loc[(df_lac[0].str.startswith('LocationArea', na=False))]
    df_lac=df_lac[0].str.split('\s+', expand=True)
    df_lac = df_lac.drop(1, axis=1)
    dict_lac=pd.Series(df_lac[2].values,index=df_lac[0]).to_dict()
    
    
    
    df_ssv_input=df_ssv_input.replace({"locationAreaRef": dict_lac})
    df_ssv_input['RAC']=1
    df_ssv_input['SAC']=df_ssv_input['cId']
    df_ssv_input['URA']="NA"
    df_ssv_input=df_ssv_input[['MO','cId','locationAreaRef','RAC','SAC','URA','uarfcnUl','uarfcnDl','primaryScramblingCode','primaryCpichPower','maximumTransmissionPower']]
    
    
    df_input_cell=df_input.loc[(df_input['Data Type'].str.startswith('Cell', na=False))]
    df_input_cell=df_input_cell.reset_index(drop=True)
    
    df_input_cell_2=pd.merge(df_input_cell, df_ssv_input, left_on='Data', right_on='MO', how='left').drop('MO', axis=1)
    df_input_cell_2 = df_input_cell_2.drop('Remarks', axis=1)
    df_input_cell_2['Data'] = df_input_cell_2['Data'].replace('NO INFO',np.nan)
    df_input_cell_2 = df_input_cell_2.drop('Data Type', axis=1)
    
    df_input_cell_3 = df_input_cell_2.transpose()
    
    df_input_cell_3 = df_input_cell_3.replace({np.nan: ""})
    
    ##################    Loading and Saving SSV INPUT  ###########################    
    
    wb = load_workbook(INPATH +os.sep+'SSV_FILE_INPUTS.xlsx')  
    
    ws = wb['QRF Input-1']
    
    rows = dataframe_to_rows(df_input_cell_3,index=False,header=False)
    for r_idx, row in enumerate(rows, 2):
    	for c_idx, value in enumerate(row, 27):
    		ws.cell(row=r_idx, column=c_idx, value=str(value))
    
    wb.save(OUTPATH+os.sep+'SSV_FILE_INPUTS.xlsx')

if typeofnode=="DU":
    # Function to get data b/w two strings

    def find_between( s, first, last ):
        try:
            start = s.index( first ) -10
            end = s.index( last, start )
            return s[start:end]
        except ValueError:
            return ""

    def find_between_rnc_commands( s, first, last ):
        try:
            start = s.index( first )
            end = s.index( last, start )
            return s[start:end]
        except ValueError:
            return ""


    def get_line_with_substring(input_string, substring):
        lines = input_string.split('\n')
        
        for line in lines:
            if substring in line:
                return line
        return None
        
    # Define border style for excel
    border = Border(
        left=Side(style="thick"),
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick")
        )

    def apply_border( cell_range ):
        # Split the cell range into top-left and bottom-right coordinates
        top_left_cell, bottom_right_cell = cell_range.split(":")
        
        # Set the outline border for the range
        min_row, min_col = ws[top_left_cell].row, ws[top_left_cell].column
        max_row, max_col = ws[bottom_right_cell].row, ws[bottom_right_cell].column
        
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if row == min_row:
                    cell.border = cell.border + Border(top=border.top)
                if row == max_row:
                    cell.border = cell.border + Border(bottom=border.bottom)
                if col == min_col:
                    cell.border = cell.border + Border(left=border.left)
                if col == max_col:
                    cell.border = cell.border + Border(right=border.right)
                
    ######################################### Baseband SSV CHECKLIST ##############################################################################

    ##################    Loading WCDMA DU  ###########################  
    print("Loading Log...")
    with open (INPATH + os.sep+ 'DU_Log.log', "r") as file:
    	data = file.read().rstrip() 

    data = data.replace("|", "_")

    ##################    Loading Workbook  ###########################    

    wb = load_workbook(INPATH +os.sep+'3G_Checklist_Template.xlsx')  

    ##################    VSWR  ########################### 

    print("Getting VSWR")    
       
    start_string="invxrf"
    end_string="st ret"
    dict_replace={'invxrf':''}

    data_between=find_between( data, start_string, end_string )

    stringio_data_between= StringIO(data_between)
    df_data_between = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")

    df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)

    ws = wb['VSWR']

    rows = dataframe_to_rows(df_data_between,index=False,header=False)
    for r_idx, row in enumerate(rows, 4):
    	for c_idx, value in enumerate(row, 2):
    		if str(value)[0]=="=":
    			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
    		else:
    			ws.cell(row=r_idx, column=c_idx, value=str(value))

    cell_range = "B4:B"+str(len(df_data_between)+3)
    apply_border( cell_range )


    ##################    RET  ########################### 

    print("Getting RET")    
       
    start_string="st ret"
    end_string="get 0"
    dict_replace={'st ret':'','lget . Electricalantennatilt':''}

    data_between=find_between( data, start_string, end_string )

    stringio_data_between= StringIO(data_between)
    df_data_between = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")

    df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)

    ws = wb['RET']

    rows = dataframe_to_rows(df_data_between,index=False,header=False)
    for r_idx, row in enumerate(rows, 4):
    	for c_idx, value in enumerate(row, 2):
    		if str(value)[0]=="=":
    			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
    		else:
    			ws.cell(row=r_idx, column=c_idx, value=str(value))

    cell_range = "B4:B"+str(len(df_data_between)+3)
    apply_border( cell_range )


    ##################    EQUIPMENT  ########################### 

    print("Getting EQUIPMENT")    
       
    start_string="get 0"
    end_string="license key"
    dict_replace={'get 0':''}

    data_between=find_between( data, start_string, end_string )

    stringio_data_between= StringIO(data_between)
    df_data_between = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")

    df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)

    ws = wb['EQUIPMENT']

    rows = dataframe_to_rows(df_data_between,index=False,header=False)
    for r_idx, row in enumerate(rows, 4):
    	for c_idx, value in enumerate(row, 2):
    		if str(value)[0]=="=":
    			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
    		else:
    			ws.cell(row=r_idx, column=c_idx, value=str(value))

    cell_range = "B4:B"+str(len(df_data_between)+3)
    apply_border( cell_range )


    ##################    BRD  ########################### 

    print("Getting BRD")    
       
    start_string="invxrf"
    end_string="st ret"
    dict_replace={'invxrf':''}

    data_between=find_between( data, start_string, end_string )

    stringio_data_between= StringIO(data_between)
    df_data_between = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")

    df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)

    ws = wb['BRD']

    rows = dataframe_to_rows(df_data_between,index=False,header=False)
    for r_idx, row in enumerate(rows, 4):
    	for c_idx, value in enumerate(row, 2):
    		if str(value)[0]=="=":
    			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
    		else:
    			ws.cell(row=r_idx, column=c_idx, value=str(value))

    cell_range = "B4:B"+str(len(df_data_between)+3)
    apply_border( cell_range )


    ##################    LICENSE  ########################### 

    print("Getting LICENSE")    
       
    start_string="license key"
    end_string="> pst"
    dict_replace={'license key':''}

    data_between=find_between( data, start_string, end_string )

    stringio_data_between= StringIO(data_between)
    df_data_between = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")

    df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)

    ws = wb['LICENSE']

    rows = dataframe_to_rows(df_data_between,index=False,header=False)
    for r_idx, row in enumerate(rows, 4):
    	for c_idx, value in enumerate(row, 2):
    		if str(value)[0]=="=":
    			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
    		else:
    			ws.cell(row=r_idx, column=c_idx, value=str(value))

    cell_range = "B4:B"+str(len(df_data_between)+3)
    apply_border( cell_range )


    ##################    COUNTER  ########################### 

    print("Getting COUNTER")    
       
    start_string="> pst"
    end_string="get iub"
    dict_replace={'pst':''}

    data_between=find_between( data, start_string, end_string )

    stringio_data_between= StringIO(data_between)
    df_data_between = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")

    df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)

    ws = wb['COUNTER']

    rows = dataframe_to_rows(df_data_between,index=False,header=False)
    for r_idx, row in enumerate(rows, 4):
    	for c_idx, value in enumerate(row, 2):
    		if str(value)[0]=="=":
    			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
    		else:
    			ws.cell(row=r_idx, column=c_idx, value=str(value))

    cell_range = "B4:B"+str(len(df_data_between)+3)
    apply_border( cell_range )



    ##################    E1  ########################### 

    print("Getting E1")    
       
    start_string="get iub"
    end_string="> alt"
    dict_replace={'get iub':''}

    data_between=find_between( data, start_string, end_string )

    stringio_data_between= StringIO(data_between)
    df_data_between = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")

    df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)

    ws = wb['E1']

    rows = dataframe_to_rows(df_data_between,index=False,header=False)
    for r_idx, row in enumerate(rows, 4):
    	for c_idx, value in enumerate(row, 2):
    		if str(value)[0]=="=":
    			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
    		else:
    			ws.cell(row=r_idx, column=c_idx, value=str(value))

    cell_range = "B4:B"+str(len(df_data_between)+3)
    apply_border( cell_range )


    ##################    ALARMS  ########################### 

    print("Getting ALARMS")    
       
    start_string="> alt"
    end_string="invl power"
    dict_replace={'alt':''}

    data_between=find_between( data, start_string, end_string )

    stringio_data_between= StringIO(data_between)
    df_data_between = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")

    df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)

    ws = wb['ALARMS']

    rows = dataframe_to_rows(df_data_between,index=False,header=False)
    for r_idx, row in enumerate(rows, 4):
    	for c_idx, value in enumerate(row, 2):
    		if str(value)[0]=="=":
    			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
    		else:
    			ws.cell(row=r_idx, column=c_idx, value=str(value))

    cell_range = "B4:B"+str(len(df_data_between)+3)
    apply_border( cell_range )



    ##################    LAC  ########################### 

    print("Getting LAC")    

    line=get_line_with_substring(data, "utrancellid_localcellid_uarfcndl_uarfcnul_locationAreaRef")

    start_string=line
    end_string="get loc lac"
    dict_replace={line[9:]:''}

    data_between=find_between_rnc_commands( data, start_string, end_string )

    stringio_data_between= StringIO(data_between)
    df_data_between = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")

    df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)

    ws = wb['LAC']

    rows = dataframe_to_rows(df_data_between,index=False,header=False)
    for r_idx, row in enumerate(rows, 4):
    	for c_idx, value in enumerate(row, 2):
    		if str(value)[0]=="=":
    			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
    		else:
    			ws.cell(row=r_idx, column=c_idx, value=str(value))

    cell_range = "B4:B"+str(len(df_data_between)+3)
    apply_border( cell_range )



    ##################    Working on NBRs  ########################### 

    print("Getting NBRs")   

    nbr_ws_dict={0:'NBRList_850A_S1',1:'NBRList_850A_S2',2:'NBRList_850A_S3',3:'NBRList_850A_S4',
                 4:'NBRList_850B_S1',5:'NBRList_850B_S2',6:'NBRList_850B_S3',7:'NBRList_850B_S4',
                 8:'NBRList_2100A_S1',9:'NBRList_2100A_S2',10:'NBRList_2100A_S3',11:'NBRList_2100A_S4',
                 12:'NBRList_2100B_S1',13:'NBRList_2100B_S2',14:'NBRList_2100B_S3',15:'NBRList_2100B_S4'}
    cell_count=0

    for cell in cell_list:
        if cell=="NO INFO":
            ws = wb[nbr_ws_dict[cell_count]]
            ws.sheet_state = 'hidden'
            cell_count=cell_count+1
        else:
            start_string="lget utrancell="+cell+' utranrelation_gsmrelation'
            end_string="MOs"
            dict_replace={start_string:''}

            data_between=find_between( data, start_string, end_string )

            stringio_data_between= StringIO(data_between)
            df_data_between = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")

            df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)

            ws = wb[nbr_ws_dict[cell_count]]

            rows = dataframe_to_rows(df_data_between,index=False,header=False)
            for r_idx, row in enumerate(rows, 4):
            	for c_idx, value in enumerate(row, 2):
            		if str(value)[0]=="=":
            			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
            		else:
            			ws.cell(row=r_idx, column=c_idx, value=str(value))

            cell_range = "B4:B"+str(len(df_data_between)+3)
            apply_border( cell_range )
            
            cell_count=cell_count+1

            
    ########### Saving 3G_Checklist ###########

    print("Saving 3G_Checklist.")  

    now = datetime.now()
    today = date.today()


    date_today = time.strftime('%b-%d-%Y', time.localtime())
    wb.save(OUTPATH+os.sep+'3G_Checklist_'+nodename+"_"+date_today+'.xlsx')



    ########### Working on power checklist ###########

    ##################    Loading Workbook  ###########################    

    wb = load_workbook(INPATH +os.sep+'3G_Checklist_Power_Template.xlsx')  


    ##################    POWER  ########################### 

    print("Getting POWER")    
       
    start_string="invl power"
    end_string="get . maxTotalOutputPower"
    dict_replace={'invl power':''}

    data_between=find_between( data, start_string, end_string )

    stringio_data_between= StringIO(data_between)
    df_data_between = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")

    df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)

    ws = wb['POWER']

    rows = dataframe_to_rows(df_data_between,index=False,header=False)
    for r_idx, row in enumerate(rows, 4):
    	for c_idx, value in enumerate(row, 2):
    		if str(value)[0]=="=":
    			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
    		else:
    			ws.cell(row=r_idx, column=c_idx, value=str(value))

    cell_range = "B4:B"+str(len(df_data_between)+3)
    apply_border( cell_range )


    ##################    MAX_POWER_CONFIG  ########################### 

    print("Getting MAX_POWER_CONFIG")    
       
    start_string="get . maxTotalOutputPower"
    end_string="exit"
    dict_replace={'get . maxTotalOutputPower':''}

    data_between=find_between( data, start_string, end_string )

    stringio_data_between= StringIO(data_between)
    df_data_between = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")

    df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)

    ws = wb['MAX_POWER_CONFIG']

    rows = dataframe_to_rows(df_data_between,index=False,header=False)
    for r_idx, row in enumerate(rows, 4):
    	for c_idx, value in enumerate(row, 2):
    		if str(value)[0]=="=":
    			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
    		else:
    			ws.cell(row=r_idx, column=c_idx, value=str(value))

    cell_range = "B4:B"+str(len(df_data_between)+3)
    apply_border( cell_range )



    ########### Saving Power CheckList ###########

    print("Saving Power CheckList.")  

    now = datetime.now()
    today = date.today()


    date_today = time.strftime('%b-%d-%Y', time.localtime())
    wb.save(OUTPATH+os.sep+'3G_Checklist_Power_'+nodename+"_"+date_today+'.xlsx')


    ######################## Working on SSV_FILE_INPUTS #######################


    print("Saving SSV_FILE_INPUTS.")

    line=get_line_with_substring(data, "cId_locationAreaRef_maximumTransmissionPower_primaryCpichPower_primaryScramblingCode_uarfccnDl_uarfcnUl")

    start_string=line
    end_string="get loc lac"
    dict_replace={line[9:]:''}
     
    data_between=find_between_rnc_commands( data, start_string, end_string )

    stringio_data_between= StringIO(data_between)
    df_ssv_input = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")

    df_ssv_input=df_ssv_input.loc[(df_ssv_input[0].str.startswith('Utran', na=False))|(df_ssv_input[0].str.startswith('MO', na=False))]

    df_ssv_input=df_ssv_input[0].str.split('\s+', expand=True)

    df_ssv_input[0] = df_ssv_input[0].str.replace('UtranCell=','')

    df_ssv_input = df_ssv_input.reset_index(drop=True)

    new_header = df_ssv_input.iloc[0] 
    df_ssv_input = df_ssv_input[1:]
    df_ssv_input.columns = new_header

    df_ssv_input = df_ssv_input.reset_index(drop=True)

    print("Mapping LAC.")

    start_string="get loc lac"
    end_string="Total"
    dict_replace={start_string:''}

    data_between=find_between( data, start_string, end_string )

    stringio_data_between= StringIO(data_between)
    df_lac = pd.read_csv(stringio_data_between, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")


    df_lac=df_lac.loc[(df_lac[0].str.startswith('LocationArea', na=False))]
    df_lac=df_lac[0].str.split('\s+', expand=True)
    df_lac = df_lac.drop(1, axis=1)
    dict_lac=pd.Series(df_lac[2].values,index=df_lac[0]).to_dict()



    df_ssv_input=df_ssv_input.replace({"locationAreaRef": dict_lac})
    df_ssv_input['RAC']=1
    df_ssv_input['SAC']=df_ssv_input['cId']
    df_ssv_input['URA']="NA"
    df_ssv_input=df_ssv_input[['MO','cId','locationAreaRef','RAC','SAC','URA','uarfcnUl','uarfcnDl','primaryScramblingCode','primaryCpichPower','maximumTransmissionPower']]


    df_input_cell=df_input.loc[(df_input['Data Type'].str.startswith('Cell', na=False))]
    df_input_cell=df_input_cell.reset_index(drop=True)

    df_input_cell_2=pd.merge(df_input_cell, df_ssv_input, left_on='Data', right_on='MO', how='left').drop('MO', axis=1)
    df_input_cell_2 = df_input_cell_2.drop('Remarks', axis=1)
    df_input_cell_2['Data'] = df_input_cell_2['Data'].replace('NO INFO',np.nan)
    df_input_cell_2 = df_input_cell_2.drop('Data Type', axis=1)

    df_input_cell_3 = df_input_cell_2.transpose()

    df_input_cell_3 = df_input_cell_3.replace({np.nan: ""})

    ##################    Loading and Saving SSV INPUT  ###########################    

    wb = load_workbook(INPATH +os.sep+'SSV_FILE_INPUTS.xlsx')  

    ws = wb['QRF Input-1']

    rows = dataframe_to_rows(df_input_cell_3,index=False,header=False)
    for r_idx, row in enumerate(rows, 2):
    	for c_idx, value in enumerate(row, 27):
    		ws.cell(row=r_idx, column=c_idx, value=str(value))

    wb.save(OUTPATH+os.sep+'SSV_FILE_INPUTS.xlsx')