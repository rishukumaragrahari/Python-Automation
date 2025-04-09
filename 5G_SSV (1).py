# -*- coding: utf-8 -*-
"""
Created on Thu Mar 24 17:24:37 2020

@author: Rishav Kumar Agrahari
"""
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from openpyxl.styles import Border, Side
import time
if sys.version_info[0] < 3: 
   	from StringIO import StringIO
else:
   	from io import StringIO

INPATH =r"C:\\Users\\ezagrri\\OneDrive - Ericsson\\python\\Claro\\SSV_Automations\\5G\\Input\\"
OUTPATH =r"C:\\Users\\ezagrri\\OneDrive - Ericsson\\python\\Claro\\SSV_Automations\\5G\\Output\\"
    	
def find_between( s, first, last ):
    try:
        start = s.index( first ) - 12
        end = s.index( last, start )
        return s[start:end]
    except ValueError:
        return ""

border = Border(
    left=Side(style="thick"),
    right=Side(style="thick"),
    top=Side(style="thick"),
    bottom=Side(style="thick")
)

def apply_border( cell_range ):
    # Split the cell range into top-left and bottom-right coordinates
    top_left_cell, bottom_right_cell = cell_range.split(":")
    
    # Set the outline border for the range
    min_row, min_col = ws[top_left_cell].row, ws[top_left_cell].column
    max_row, max_col = ws[bottom_right_cell].row, ws[bottom_right_cell].column
    
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if row == min_row:
                cell.border = cell.border + Border(top=border.top)
            if row == max_row:
                cell.border = cell.border + Border(bottom=border.bottom)
            if col == min_col:
                cell.border = cell.border + Border(left=border.left)
            if col == max_col:
                cell.border = cell.border + Border(right=border.right)

df_input = pd.read_excel(INPATH+os.sep+'Input_Details.xlsx', sheet_name="Input")
typeofradio=df_input.at[0,'Data']

print("Loading Log...")
with open (INPATH + os.sep+ '5G_SSV.log', "r") as file:
	data = file.read().rstrip()



##################    Loading Workbook  ###########################    

wb = load_workbook(INPATH +os.sep+'5G_TDD_Checklist_Template.xlsx')    

##################    SITE_STATUS  ########################### 

print("Getting SITE_STATUS")    
   
start_string="st sector"
end_string="hget NRCellDU cellLocalId|nRPCI|rachPreambleFormat|rachRootSequence"
dict_replace={'st sector':'', 'get NRSectorCarrier arfcndl':'','get NRSectorCarrier bsChannelBwDl':'', 'hget nrcelldu bandlist':''}

data_between=find_between( data, start_string, end_string )

SITE_STATUS= StringIO(data_between)
df_data_between = pd.read_csv(SITE_STATUS, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")

df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)

ws = wb['SITE_STATUS']

rows = dataframe_to_rows(df_data_between,index=False,header=False)
for r_idx, row in enumerate(rows, 4):
	for c_idx, value in enumerate(row, 2):
		if str(value)[0]=="=":
			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
		else:
			ws.cell(row=r_idx, column=c_idx, value=str(value))

cell_range = "B4:B"+str(len(df_data_between)+3)
apply_border( cell_range )


##################    ID  ########################### 

print("Getting ID")    
   
start_string="hget NRCellDU cellLocalId|nRPCI|rachPreambleFormat|rachRootSequence"
end_string="get 0"
dict_replace={'hget NRCellDU cellLocalId|nRPCI|rachPreambleFormat|rachRootSequence':'','get NRCellDU tddUlDlPattern':''}

data_between=find_between( data, start_string, end_string )

SITE_STATUS= StringIO(data_between)
df_data_between = pd.read_csv(SITE_STATUS, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")

df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)

ws = wb['ID']

rows = dataframe_to_rows(df_data_between,index=False,header=False)
for r_idx, row in enumerate(rows, 4):
	for c_idx, value in enumerate(row, 2):
		if str(value)[0]=="=":
			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
		else:
			ws.cell(row=r_idx, column=c_idx, value=str(value))

cell_range = "B4:B"+str(len(df_data_between)+3)
apply_border( cell_range )


##################    SW  ########################### 


print("Getting SW")    
   
start_string="get 0"
end_string="st AIR"
dict_replace={'get 0':''}

data_between=find_between( data, start_string, end_string )

SITE_STATUS= StringIO(data_between)
df_data_between = pd.read_csv(SITE_STATUS, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")

df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)

ws = wb['SW']

rows = dataframe_to_rows(df_data_between,index=False,header=False)
for r_idx, row in enumerate(rows, 4):
	for c_idx, value in enumerate(row, 2):
		if str(value)[0]=="=":
			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
		else:
			ws.cell(row=r_idx, column=c_idx, value=str(value))

cell_range = "B4:B"+str(len(df_data_between)+3)
apply_border( cell_range )




##################    HW  ########################### 

print("Getting HW")    

if typeofradio=="AIR":   
    start_string="st AIR"
    end_string="st rru"
    dict_replace={'st AIR':'','invhr':''}

if typeofradio=="RRU": 
    start_string="st rru"
    end_string="get . digitalTilt"
    dict_replace={'st rru':'','invxrf':''}

if typeofradio=="BOTH": 
    start_string="st AIR"
    end_string="get . digitalTilt"
    dict_replace={'st AIR':'','invhr':'','st rru':'','invxrf':''}
    

data_between=find_between( data, start_string, end_string )

SITE_STATUS= StringIO(data_between)
df_data_between = pd.read_csv(SITE_STATUS, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")

df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)

ws = wb['HW']

rows = dataframe_to_rows(df_data_between,index=False,header=False)
for r_idx, row in enumerate(rows, 4):
	for c_idx, value in enumerate(row, 2):
		if str(value)[0]=="=":
			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
		else:
			ws.cell(row=r_idx, column=c_idx, value=str(value))

cell_range = "B4:B"+str(len(df_data_between)+3)
apply_border( cell_range )



##################    RET  ########################### 

print("Getting RET")    

if typeofradio=="AIR":   
    start_string="get . digitalTilt"
    end_string="get . electricalAntennaTilt"
    dict_replace={'get . digitalTilt':''}

if typeofradio=="RRU": 
    start_string="get . electricalAntennaTilt"
    end_string="alt"
    dict_replace={'get . electricalAntennaTilt':''}

if typeofradio=="BOTH": 
    start_string="get . digitalTilt"
    end_string="alt"
    dict_replace={'get . digitalTilt':'','get . electricalAntennaTilt':''}
    

data_between=find_between( data, start_string, end_string )

SITE_STATUS= StringIO(data_between)
df_data_between = pd.read_csv(SITE_STATUS, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")

df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)

ws = wb['RET']

rows = dataframe_to_rows(df_data_between,index=False,header=False)
for r_idx, row in enumerate(rows, 4):
	for c_idx, value in enumerate(row, 2):
		if str(value)[0]=="=":
			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
		else:
			ws.cell(row=r_idx, column=c_idx, value=str(value))

cell_range = "B4:B"+str(len(df_data_between)+3)
apply_border( cell_range )




##################    ALARMS  ########################### 


print("Getting ALARMS")    
   
start_string="alt"
end_string="get NRSectorCarrier configuredMaxTxPower"
dict_replace={'alt':''}

data_between=find_between( data, start_string, end_string )

SITE_STATUS= StringIO(data_between)
df_data_between = pd.read_csv(SITE_STATUS, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")

df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)

ws = wb['ALARMS']

rows = dataframe_to_rows(df_data_between,index=False,header=False)
for r_idx, row in enumerate(rows, 4):
	for c_idx, value in enumerate(row, 2):
		if str(value)[0]=="=":
			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
		else:
			ws.cell(row=r_idx, column=c_idx, value=str(value))

cell_range = "B4:B"+str(len(df_data_between)+3)
apply_border( cell_range )


##################    Power  ########################### 


print("Getting Power")    
   
start_string="get NRSectorCarrier configuredMaxTxPower"
end_string="#END OF SCRIPT"
dict_replace={'get NRSectorCarrier configuredMaxTxPower':''}

data_between=find_between( data, start_string, end_string )

SITE_STATUS= StringIO(data_between)
df_data_between = pd.read_csv(SITE_STATUS, engine="python", index_col=False, header = None , on_bad_lines='skip', sep=r"\n")

df_data_between[0] = df_data_between[0].replace(dict_replace,regex=True)

ws = wb['Power']

rows = dataframe_to_rows(df_data_between,index=False,header=False)
for r_idx, row in enumerate(rows, 4):
	for c_idx, value in enumerate(row, 2):
		if str(value)[0]=="=":
			ws.cell(row=r_idx, column=c_idx, value="'"+str(value))
		else:
			ws.cell(row=r_idx, column=c_idx, value=str(value))

cell_range = "B4:B"+str(len(df_data_between)+3)
apply_border( cell_range )


##################    END OF CHECKLIST  ########################### 

##################    Saving CHECKLIST  ########################### 

nodename=df_data_between[0][0][:10]
date_today = time.strftime('%b-%d-%Y', time.localtime())


wb.save(OUTPATH+os.sep+'5G_Checklist_'+nodename+"_"+date_today+".xlsx")